from TaskboardManager import TaskboardManager
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Body
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from io import BytesIO
from fastapi.responses import FileResponse
from typing import Optional, Annotated

app = FastAPI(title="Taskboard")
app.mount("/static", StaticFiles(directory="static",
          html=True), name="static")

# Enable CORS (Cross-Origin Resource Sharing) for multi-user interfaces (automated handling of synchronous clients.)
origins = ["*"]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Create an instance of TaskboardManager to manage Taskboards
db = TaskboardManager()


@app.get("/")
async def home():
    return FileResponse("static/index.html")


@app.get("/settings")
async def settings():
    return FileResponse("static/settings.html")


@app.get("/edit")
async def edit():
    return FileResponse("static/edit.html")


@app.get("/menu")
async def menu():
    return FileResponse("static/menu.html")


@app.post("/create")
async def create_taskboard(name: Annotated[str, Form()], file: UploadFile = File()):
    db.createTaskboard(name)

    contents = await file.read()
    if contents != b"":
        # Process the file into the required format
        workbook = load_workbook(filename=BytesIO(contents))
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        headers = rows[0]

        data = [dict(zip(headers, row))
                for row in rows[1:]]  # Convert rows to list of dicts

        db.uploadFile(name, data)
        return {"msg": "file uploaded and data processed"}
    return {"msg": "file not uploaded"}


@app.post("/duplicate")
async def duplicate_taskboard(currentName: str, newName: str):
    try:
        # Check if the new taskboard name already exists
        if newName in db.listAvailable():
            return {"message": "Taskboard with the new name already exists."}

        # Retrieve data from the current taskboard
        currentData = db.getAsDict(currentName)

        # Convert the data into a list of dictionaries if it's not already
        if isinstance(currentData, dict):
            currentData = list(currentData.values())

        # Ensure the data is in the correct format: a list of dictionaries
        if not isinstance(currentData, list) or not all(isinstance(item, dict) for item in currentData):
            raise ValueError(
                "Invalid data format: Data must be a list of dictionaries.")

        # Create a new taskboard and upload the data
        db.createTaskboard(newName)
        db.uploadFile(newName, currentData)
        return {"message": "Taskboard duplicated successfully."}
    except Exception as e:
        print(f"Error during duplication: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/upload/")
async def upload_file(name: str = Form(...), file: UploadFile = File(...)):
    contents = await file.read()
    workbook = load_workbook(filename=BytesIO(contents))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = rows[0]

    data = [dict(zip(headers, row))
            for row in rows[1:]]  # Convert rows to list of dicts
    db.uploadFile(name, data)  # Call the updated uploadFile method
    return {"message": f"Data uploaded successfully to taskboard '{name}'."}


@app.delete("/delete/")
async def delete_taskboard(name: str):
    db.destroy(name)
    return


@app.put("/rollback/")
async def rollback_taskboard(name: str, rollback: Optional[int] = 1):
    db.rollback(name, rollback)
    return


@app.get("/list/")
async def list_taskboards():
    return db.defAvailable()


@app.get("/get/")
async def get_taskboard(name: str):
    return db.getAsDict(name)


@app.get("/get-settings/")
async def get_taskboard_settings(name: str):
    try:
        # Retrieve the current settings for the taskboard
        settings = db.getSettings(name)
        return settings
    except Exception as e:
        return {"error": str(e)}

@app.post("/update-task")
async def update_task(update_details: dict = Body(...)):
    taskboardName = update_details['taskboardName']
    taskId = update_details['taskId']
    key = update_details['key']
    value = update_details['value']

    try:
        db.updateTask(taskboardName, taskId, key, value)  # Pass font size to your TaskboardManager
        return {"message": "Task updated successfully."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/save-font-size")
async def save_font_size(name: str, font_size: str):
    try:
        db.saveFontsize(name, font_size)
        return {"message": "Font size saved successfully."}
    except Exception as e:
        return {"error": str(e)}





@app.post("/save-bg-color")
async def save_bg_color(name: str, bg_color: str):
    try:
        db.saveBgColor(name, bg_color)
        return {"message": "Background color saved successfully."}
    except Exception as e:
        return {"error": str(e)}
    
    
@app.post("/save-font-color")
async def save_font_color(name: str, font_color: str):
    try:
        db.saveFontColor(name, font_color)
        return {"message": "Background color saved successfully."}
    except Exception as e:
        return {"error": str(e)}
    
    

@app.post("/save-view-headers")
async def save_view_headers(name: str, view_header: str):
    try:
        db.saveViewHeader(name, view_header)
        return {"message": "Background color saved successfully."}
    except Exception as e:
        return {"error": str(e)}




@app.post("/save-title-header")
async def save_title_header(name: str, title_header: str):
    try:
        db.saveTitleHeader(name, title_header)
        return {"message": "Background color saved successfully."}
    except Exception as e:
        return {"error": str(e)}
    
    
    
    
    
@app.post("/save-text-wrapping")
async def save_text_wrapping(name: str, text_wrapping: str):
    try:
        db.saveTextWrapping(name, text_wrapping)
        return {"message": "Text-Wrapping saved successfully."}
    except Exception as e:
        return {"error": str(e)}

    
    
    
    
    

@app.post("/save-sorting-header")
async def save_sorting_header(name: str, sorting_header: str):
    try:
        db.saveSortingHeader(name, sorting_header)
        return {"message": "Background color saved successfully."}
    except Exception as e:
        return {"error": str(e)}



@app.post("/save-sort-by")
async def save_sort_by(name: str, sort_by: str):
    try:
        db.saveSortBy(name, sort_by)
        return {"message": "Background color saved successfully."}
    except Exception as e:
        return {"error": str(e)}



if __name__ == "__main__":
    import uvicorn

    print(db.listAvailable())

    if "My New Taskboard" not in db.listAvailable():
        # opening for reading as binary
        in_file = open("resources/Book2.xlsx", "rb")
        db.createTaskboard("My New Taskboard")
        db.uploadFile("My New Taskboard", in_file.read())
        in_file.close()

    uvicorn.run("app:app", host="localhost", port=8001, reload=True)
